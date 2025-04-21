from openpyxl import workbook, load_workbook
from openpyxl.utils import get_column_letter

WB = load_workbook("BSIT.xlsx")
WS = WB.active  # Get the active sheet

# Accessing multiple cells
for row in range (1,11):
    for col in range (1,5):
        char = get_column_letter(col)
        print(WS[char+ str(row)].value)
        #WS[char+ str(row)] = char + str(row)
        # print(WS.cell(row=row, column=col).value) # Another way to access cells

WB.save("BSIT.xlsx")
print("Excel file updated successfully!")
