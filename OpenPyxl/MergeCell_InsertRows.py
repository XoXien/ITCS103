from openpyxl import load_workbook, Workbook
from openpyxl.styles import Alignment
from openpyxl.styles import Font

# Load the existing Excel workbook
WB = load_workbook("Grades.xlsx")
WS = WB.active  

# Merge cells value
WS.merge_cells('A4:E4')  # Merge cells A4 to E4
WS['A4'] = 'Grades'  # Set the value of the merged cell
WS['A4'].alignment = Alignment(horizontal='center', vertical='center')  # Center align the merged cell
WS['A4'].font = Font(bold=True, size=12)  # Make the font bold and set size

WB.save("Grades.xlsx")
print("Excel file updated successfully!")