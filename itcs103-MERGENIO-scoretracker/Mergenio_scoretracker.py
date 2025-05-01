import openpyxl
from openpyxl import Workbook, load_workbook
import os
import tkinter as tk
from tkinter import messagebox, ttk

# Constants
FILE_NAME = "student_score.xlsx"

def check_wb():
    """Initialize the workbook and create a sheet if it doesn't exist."""
    if not os.path.exists(FILE_NAME):
        wb = Workbook()
        ws = wb.active
        ws.title = "Scores"
        ws.append(["Name", "Score", "Status"])  # Column headers
        wb.save(FILE_NAME)

def add_update(name, score):
    """Add a new student or update an existing student's score."""
    check_wb()
    wb = load_workbook(FILE_NAME)
    ws = wb.active

    updated = False
    for row in ws.iter_rows(min_row=2, values_only=False):
        if row[0].value == name:
            row[1].value = score
            row[2].value = status(score)
            updated = True
            break

    if not updated:
        ws.append([name, score, status(score)])

    wb.save(FILE_NAME)
    return updated

def get_records():
    """Retrieve all student records from the workbook."""
    check_wb()
    wb = load_workbook(FILE_NAME)
    ws = wb.active

    records = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        records.append(row)
    return records

def status(score):
    """Determine the status based on the student's score."""
    if score >= 75:
        return "Pass"
    else:
        return "Fail"

# === GUI commands ===
def submit_score():
    """Handle the submission of a student's score."""
    name = name_entry.get().strip()
    try:
        score = int(score_entry.get().strip())
        if not (0 <= score <= 100):
            raise ValueError
    except ValueError:
        messagebox.showerror("Invalid Input", "Score must be an integer between 0 and 100.")
        return

    if name == "":
        messagebox.showerror("Invalid Input", "Name cannot be empty.")
        return

    updated = add_update(name, score)

    if updated:
        messagebox.showinfo("Success", f"Record updated for {name}.")
    else:
        messagebox.showinfo("Success", f"Record added for {name}.")

    name_entry.delete(0, tk.END)
    score_entry.delete(0, tk.END)
    refresh_table()

def refresh_table():
    """Refresh the displayed data in the Treeview table."""
    for item in tree.get_children():
        tree.delete(item)

    records = get_records()
    for record in records:
        tree.insert("", tk.END, values=record)


# === GUI Setup ===
window = tk.Tk()
window.title("Mergenio Score Tracker")
window.geometry("370x370")
window.resizable(True, True)


# === Input Section ===
frame = tk.Frame(window)
frame.pack(pady=10)

sn_label = tk.Label(frame, text="Student Name:")
score_label = tk.Label(frame, text="Score:")


name_entry = tk.Entry(frame)
name_entry.insert(0, "Enter Name")  

score_entry = tk.Entry(frame)
score_entry.insert(0, "0 - 100")

submit_score_btn = tk.Button(frame, text="Submit Score", command=submit_score)

score_label.grid(row=1, column=0, padx=5, pady=5)
sn_label.grid(row=0, column=0, padx=5, pady=5)
name_entry.grid(row=0, column=1, padx=5, pady=5)
score_entry.grid(row=1, column=1, padx=5, pady=5)
submit_score_btn.grid(row=2, column=0, columnspan=2, pady=10)

# === Table Section ===
frame = tk.Frame(window)
frame.pack(pady=10)

tree = ttk.Treeview(frame, columns=("Name", "Score", "Status"), show="headings", height=10)

tree.heading("Name", text="Name")
tree.heading("Score", text="Score")
tree.heading("Status", text="Status")

tree.column("Name", width=100)
tree.column("Score", width=100)
tree.column("Status", width=100)
tree.pack()

#Run the application
check_wb()
refresh_table()
window.mainloop()