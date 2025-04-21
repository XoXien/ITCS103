from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter

WB = load_workbook("BSIT.xlsx")
WS = WB.active  

#Insert Row and Column
WS.insert_rows(2)  
WS.insert_cols(2)

#Delete Row and Column
WS.delete_rows(2)
WS.delete_cols(2)

#Move Row and Column
WS.move_range("A1:D1", rows=2, cols=2)

WB.save("BSIT.xlsx")
print("Excel file updated successfully!")